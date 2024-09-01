import zipfile
import openpyxl
import os
import warnings

warnings.simplefilter('ignore')

SRC_TITLE_ROW = 4
DEST_TITLE_ROW = 4
DEST_START_ROW = 8
DEST_TITLE_COL_START = 29
DEST_TITLE_COL_END = [79, 77]
SRC_START_ROW = 7
DEST_TITLE_DATE_COL = [[28, 33, 38, 43, 45, 51, 53, 58, 65, 77], [29, 38, 45, 53, 59, 65, 72]]

DEST_FILE = ['Vessel.xlsx', 'Shipping.xlsx']
SRC_FILE = 'DataFile'
RESULT_FILE = ['VesselResult.xlsx', 'ShippingResult.xlsx']
SHEET_NAME = ['船舶周报', '航运周报']


def main():
    excel_list = os.listdir(SRC_FILE)

    for data_type in [1]:
        print(f"Processing {DEST_FILE[data_type]}")
        dest_sheet = openpyxl.load_workbook(DEST_FILE[data_type])[SHEET_NAME[data_type]]
        title_col = {}
        title_data_col = {}

        for i in range(DEST_TITLE_COL_START, DEST_TITLE_COL_END[data_type] + 1):
            if dest_sheet.cell(row=DEST_TITLE_ROW, column=i).value is not None:
                title_col[dest_sheet.cell(row=DEST_TITLE_ROW, column=i).value] = i
                for j in range(len(DEST_TITLE_DATE_COL[data_type]) - 1, -1, -1):
                    if DEST_TITLE_DATE_COL[data_type][j] < i:
                        title_data_col[dest_sheet.cell(row=DEST_TITLE_ROW, column=i).value] = DEST_TITLE_DATE_COL[data_type][j]
                        break

        for excel in excel_list:
            if not excel.endswith('.xlsx'):
                continue
            try:
                workbook = openpyxl.load_workbook(os.path.join(SRC_FILE, excel))
            except zipfile.BadZipfile:
                continue
            sheet = workbook.active

            i = 2
            while sheet.cell(row=SRC_TITLE_ROW, column=i).value is not None:
                current_title = sheet.cell(row=SRC_TITLE_ROW, column=i).value

                if current_title not in title_col:
                    print(f"\tNot Found:\t\t{current_title}!")
                    i += 1
                    continue
                print(f"\tProcessing:\t\t{current_title}!")
                dest_col = title_col[current_title]
                dest_date_col = title_data_col[current_title]

                data = []
                j = SRC_START_ROW
                while not sheet.cell(row=j, column=1).value is None:
                    date_value = sheet.cell(row=j, column=1).value
                    col_value = sheet.cell(row=j, column=i).value
                    data.append([date_value, col_value])
                    j += 1

                data.sort(key=lambda x: x[0], reverse=True)

                for k, item in enumerate(data):
                    dest_sheet.cell(row=DEST_START_ROW + k, column=dest_date_col).value = item[0]
                    dest_sheet.cell(row=DEST_START_ROW + k, column=dest_col).value = item[1]
                    dest_sheet.cell(row=DEST_START_ROW + k, column=dest_col).number_format = '#,##0.00'
                i += 1

        dest_sheet.parent.save(RESULT_FILE[data_type])


if __name__ == '__main__':
    # sheet = openpyxl.load_workbook('test.xlsx').active
    # sheet['A1'].value = 2
    # sheet.parent.save('test2.xlsx')
    # exit()
    main()
